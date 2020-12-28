import openpyxl
import numpy as np
from scipy.stats import norm


def compute_pnl(periods,path):
    dt = time_to_expiration * (1.0 - (periods/num_of_periods))
    d1 = call_d1(path[:-1], strike, risk_free_rate, vol, dt[:-1])
    d2 = call_d2(d1, vol, dt[:-1])
    portfolio_value = np.zeros(num_of_periods + 1)
    portfolio_value[0] = price(path[0], d1[0], strike, d2[0], risk_free_rate, dt[0]) * num_of_options
    delta = call_delta(d1)
    stocks_held = num_of_options * delta
    cash_account = np.zeros(num_of_periods + 1)
    cash_account[0] = portfolio_value[0] - (stocks_held[0] * path[0])
    for p in periods[1:]:
        portfolio_value[p] = stocks_held[p - 1] * path[p] + (cash_account[p - 1] * (1 + risk_free_rate / 200.0))
        if p != 50:
            cash_account[p] = portfolio_value[p] - (stocks_held[p] * path[p])
    total_hedging_pnl = portfolio_value[num_of_periods] - (num_of_options * max(0.0, path[-1] - strike))
    return total_hedging_pnl

def get_annual_real_vol(path):
    log_returns = np.array([np.log(i/j) for i, j in zip(path[1:], path)])
    vol = np.std(log_returns,ddof=1) * np.sqrt(4.0 * 50.0)
    return vol

def call_d1( asset_price, strike_price, risk_free_rate, volatility, dt):
    return (np.log((asset_price / strike_price)) + (risk_free_rate + volatility**2 / 2) * dt) / (
                volatility * np.sqrt(dt))

def call_d2(d1, volatility, dt):
        return d1 - (volatility * np.sqrt(dt))


def price(asset_price, d1, strike_price, d2, risk_free_rate, dt):
    n1 = norm.cdf(d1)
    n2 = norm.cdf(d2)
    return asset_price * n1 - strike_price * (np.exp(-(risk_free_rate * dt))) * n2

def call_delta(d1):
        return norm.cdf(d1)


time_to_expiration = 0.25
num_of_periods = 50
risk_free_rate = 0.02
num_of_options = 100000.0
vol = 0.3
strike = 50.0

wb = openpyxl.load_workbook("EquityDerivsPractice_PSet3.xlsx", read_only=True, data_only=True)
ws = wb["StockPricePaths"]
periods = np.array([[i.value for i in j] for j in ws['A2':'A52']]).flatten()

#ws = wb["DemoSheet"]
#periods = np.array([[i.value for i in j] for j in ws['D2':'D52']]).flatten()
#path1 = np.array([[i.value for i in j] for j in ws['E2':'E52']]).flatten()

path1 = np.array([[i.value for i in j] for j in ws['B2':'B52']]).flatten()
print("\nQ1:Compute the annualized realized volatility of the log-returns for price path #1")
realized_vol = get_annual_real_vol(path1)
print("\nAnswer:{}".format(round(realized_vol*100.0,2)))
print("\nQ2:Compute the total realized P&L when hedging using price path #1")
pnl = compute_pnl(periods,path1)
print("\nAnswer:{}".format(round(pnl,0)))

path2 = np.array([[i.value for i in j] for j in ws['C2':'C52']]).flatten()
print("\nQ3:Compute the annualized realized volatility of the log-returns for price path #2")
realized_vol = get_annual_real_vol(path2)
print("\nAnswer:{}".format(round(realized_vol*100.0,2)))
print("\nQ4:Compute the total realized P&L when hedging using price path #2")
pnl = compute_pnl(periods,path2)
print("\nAnswer:{}".format(round(pnl,0)))

path3 = np.array([[i.value for i in j] for j in ws['D2':'D52']]).flatten()
print("\nQ5:Compute the annualized realized volatility of the log-returns for price path #3")
realized_vol = get_annual_real_vol(path3)
print("\nAnswer:{}".format(round(realized_vol*100.0,2)))
print("\nQ6:Compute the total realized P&L when hedging using price path #3")
pnl = compute_pnl(periods,path3)
print("\nAnswer:{}".format(round(pnl,0)))

path4 = np.array([[i.value for i in j] for j in ws['E2':'E52']]).flatten()
print("\nQ7:Compute the annualized realized volatility of the log-returns for price path #4")
realized_vol = get_annual_real_vol(path4)
print("\nAnswer:{}".format(round(realized_vol*100.0,2)))
print("\nQ8:Compute the total realized P&L when hedging using price path #4")
pnl = compute_pnl(periods,path4)
print("\nAnswer:{}".format(round(pnl,0)))





